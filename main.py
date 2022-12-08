#! python3

import os
import subprocess
import openpyxl

#TODO: row 53: change A1 and B1 to actual cell values (corrupts file if changed with {}.format)

def createFolder(path):
    ''' creates folder if it doesn't already exist '''
    try:
        if not os.path.exists(path):
            os.mkdir(path)
    except OSError:
        print('Error creating directory' + path)


def open_excel(excel_file_path):
    try:
        wb = openpyxl.load_workbook(excel_file_path, data_only=True) # , data_only=True because scheduling file has a lot of formulas
        ws = wb.active
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        ws = wb.active
    wb.save(excel_file_path)
    return wb, ws


createFolder('.\\AUMID Excel')
wb, ws = open_excel('.\\AUMID Excel\AUMID Excel.xlsx')
startapps = subprocess.run('"C:\\Windows\\System32\\WindowsPowerShell\\v1.0\\powershell.exe" get-StartApps',capture_output=True, text=True, shell=True)
print(startapps)
print(startapps.stdout)
list_of_startapps = (startapps.stdout).split('  ')
print(list_of_startapps)
list_of_startapps = list(filter(lambda a: a != '', list_of_startapps))
print(list_of_startapps)
        
name = []
appID = []
for i in list_of_startapps:
    if list_of_startapps.index(i)%2 == 0:
        name.append(i)
    else:
        appID.append(i)
print(name)
print(appID)

for i in name:
    ws.cell(row=name.index(i)+1, column=1).value = i
for i in appID:
    ws.cell(row=appID.index(i)+1, column=2).value = i
for i in range(1, ws.max_row+1):
    ws.cell(row=i, column=3).value = '=_xlfn.CONCAT(CHAR(34),A1,CHAR(34),CHAR(58)," ",CHAR(34),B1,CHAR(34),CHAR(44))' #.format(ws.cell(row=i, column=1).value, ws.cell(row=i, column=2).value)
wb.save('.\\AUMID Excel\AUMID Excel.xlsx')
